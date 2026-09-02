#!/usr/bin/env python3
"""Konvertiert ein rekursives Windows-dir-Listing in eine Excel-Dateiliste.

Benötigt Python 3 und openpyxl:
    python -m pip install openpyxl

Beispiel:
    python dir_listing_to_excel.py dir.txt dateiliste.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path, PureWindowsPath

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit(
        "Das Paket 'openpyxl' fehlt. Installation: python -m pip install openpyxl"
    ) from exc


DIRECTORY_RE = re.compile(r"^\s*(?:Verzeichnis von|Directory of)\s+(.+?)\s*$", re.I)
FILE_RE = re.compile(
    r"^\s*(\d{2}[./-]\d{2}[./-]\d{4})\s+"
    r"(\d{2}:\d{2})(?:\s+[AP]M)?\s+"
    r"([\d.,]+)\s+(.+?)\s*$",
    re.I,
)


def read_listing(path: Path, encoding: str | None = None) -> str:
    """Liest typische Ausgaben von `dir /s`, einschließlich OEM-Codepage 850."""
    data = path.read_bytes()
    if encoding:
        return data.decode(encoding)

    if data.startswith((b"\xff\xfe", b"\xfe\xff")) or b"\x00" in data[:200]:
        for candidate in ("utf-16", "utf-16-le"):
            try:
                return data.decode(candidate)
            except UnicodeDecodeError:
                pass

    for candidate in ("utf-8-sig", "cp850", "cp1252", "latin-1"):
        try:
            return data.decode(candidate)
        except UnicodeDecodeError:
            pass
    raise UnicodeError("Die Zeichencodierung des Listings konnte nicht erkannt werden.")


def parse_listing(text: str, mb_base: int) -> list[tuple[datetime, float, str, str]]:
    """Extrahiert ausschließlich Dateien; Verzeichnisse und Summen werden ignoriert."""
    current_directory: str | None = None
    records: list[tuple[datetime, float, str, str]] = []

    for raw_line in text.splitlines():
        directory_match = DIRECTORY_RE.match(raw_line)
        if directory_match:
            current_directory = directory_match.group(1)
            continue

        if not current_directory or "<DIR>" in raw_line.upper():
            continue

        file_match = FILE_RE.match(raw_line)
        if not file_match:
            continue

        date_text, time_text, size_text, name = file_match.groups()
        # Summenzeilen besitzen keinen Zeitstempel und erreichen diesen Punkt nicht.
        normalized_date = date_text.replace("/", ".").replace("-", ".")
        timestamp = datetime.strptime(f"{normalized_date} {time_text}", "%d.%m.%Y %H:%M")
        size_bytes = int(re.sub(r"[.,]", "", size_text))
        folder = str(PureWindowsPath(current_directory))
        records.append((timestamp, size_bytes / mb_base, name, folder))

    return records


def write_excel(records: list[tuple[datetime, float, str, str]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dateiliste"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"

    headers = ["Datum", "Größe (MB)", "Name", "Ordnerpfad"]
    sheet.append(headers)
    for record in records:
        sheet.append(record)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22

    for cell in sheet["A"][1:]:
        cell.number_format = "dd.mm.yyyy hh:mm"
    for cell in sheet["B"][1:]:
        cell.number_format = "0.000"

    sheet.column_dimensions["A"].width = 19
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 58
    sheet.column_dimensions["D"].width = 95

    if records:
        table = Table(displayName="Dateiliste", ref=f"A1:D{len(records) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.auto_filter.ref = "A1:D1"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Erstellt aus einem rekursiven Windows-dir-Listing eine Excel-Dateiliste."
    )
    parser.add_argument("input", type=Path, help="Pfad zur dir-Textdatei")
    parser.add_argument(
        "output",
        nargs="?",
        type=Path,
        help="Zieldatei; Standard: <Eingabename>_dateiliste.xlsx",
    )
    parser.add_argument(
        "--encoding",
        help="Optionale Zeichencodierung, z. B. cp850, cp1252, utf-8 oder utf-16",
    )
    parser.add_argument(
        "--mb-base",
        type=int,
        choices=(1000, 1024),
        default=1024,
        help="Umrechnungsbasis: 1024 (Standard) oder 1000",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    output = args.output or args.input.with_name(f"{args.input.stem}_dateiliste.xlsx")

    try:
        text = read_listing(args.input, args.encoding)
        records = parse_listing(text, args.mb_base**2)
        write_excel(records, output)
    except (OSError, UnicodeError, ValueError) as exc:
        print(f"Fehler: {exc}", file=sys.stderr)
        return 1

    print(f"{len(records)} Dateien nach '{output}' geschrieben.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
