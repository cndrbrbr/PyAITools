#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Outlook-PDF-Portfolios in Excel umwandeln.

Version 2.2 - kompatible Portfolio-Erkennung auch fuer aeltere pypdf-Versionen.

Unterstuetzt:
- normale PDF-Dateien mit Outlook-E-Mails
- Adobe-PDF-Portfolios mit eingebetteten Einzel-PDFs
- Mailadressen aus mailto:-Links in den eingebetteten PDFs

Installation:
    py -m pip install pypdf openpyxl
"""

from __future__ import annotations

import re
import sys
from io import BytesIO
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

COLUMNS = [
    "Datum", "Uhrzeit", "Absender", "Empfaenger", "Betreff",
    "Nachrichtentext", "Absendemailadresse", "Empfaengermailadresse"
]

HEADER_RE = re.compile(
    r"^(Von|From|An|To|Cc|Betreff|Subject|Datum|Date|Gesendet|Sent)\s*:\s*(.*)$",
    re.IGNORECASE,
)
EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.IGNORECASE)
DATE_NUMERIC_RE = re.compile(r"\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2})\b")
TIME_RE = re.compile(r"\b(\d{1,2}:\d{2}(?::\d{2})?)\b")

MONTHS = {
    "januar": 1, "februar": 2, "maerz": 3, "maerz": 3, "märz": 3,
    "april": 4, "mai": 5, "juni": 6, "juli": 7, "august": 8,
    "september": 9, "oktober": 10, "november": 11, "dezember": 12,
    "january": 1, "february": 2, "march": 3, "may": 5, "june": 6,
    "july": 7, "october": 10, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7,
    "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def clean(value: str) -> str:
    value = (value or "").replace("\x00", "").replace("\r\n", "\n").replace("\r", "\n")
    value = value.replace("\u00a0", " ").replace("\u200b", "")
    value = re.sub(r"(?<=\w)-\n(?=\w)", "", value)
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n[ \t]+", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def split_people(value: str) -> list[str]:
    return [x.strip() for x in re.split(r"\s*;\s*", value or "") if x.strip()]


def normalize_date(value: str) -> tuple[str, str]:
    value = clean(value).replace(",", " ")
    time_match = TIME_RE.search(value)
    time_text = time_match.group(1) if time_match else ""

    numeric = DATE_NUMERIC_RE.search(value)
    if numeric:
        raw = numeric.group(1)
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", raw):
            y, m, d = map(int, raw.split("-"))
        else:
            d, m, y = map(int, re.split(r"[./-]", raw))
            if y < 100:
                y += 2000 if y < 70 else 1900
        return f"{d:02d}.{m:02d}.{y:04d}", time_text

    named = re.search(
        r"\b(\d{1,2})[. ]+([A-Za-zÄÖÜäöü]+)[ ]+(\d{4})\b", value, re.IGNORECASE
    )
    if named:
        d, month_name, y = named.groups()
        month = MONTHS.get(month_name.lower())
        if month:
            return f"{int(d):02d}.{month:02d}.{int(y):04d}", time_text
    return "", time_text


def extract_mailto(reader: PdfReader) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for page in reader.pages:
        annots = page.get("/Annots")
        if not annots:
            continue
        try:
            annots = annots.get_object()
        except Exception:
            pass
        for ref in annots:
            try:
                obj = ref.get_object()
                action = obj.get("/A") or {}
                uri = str(action.get("/URI", ""))
                if uri.lower().startswith("mailto:"):
                    address = uri[7:].split("?", 1)[0].strip()
                    key = address.lower()
                    if address and key not in seen:
                        seen.add(key)
                        result.append(address)
            except Exception:
                continue
    return result


def extract_text(reader: PdfReader) -> str:
    return clean("\n".join((page.extract_text() or "") for page in reader.pages))


def parse_email(reader: PdfReader, source_name: str) -> dict[str, str] | None:
    text = extract_text(reader)
    if not text:
        return None

    # Die Deckseite eines Adobe-PDF-Portfolios ist keine E-Mail.
    lower_text = text.lower()
    if (
        "zur optimalen anzeige dieses pdf-portfolios" in lower_text
        or "adobe reader jetzt herunterladen" in lower_text
    ):
        return None

    lines = text.splitlines()
    headers: dict[str, str] = {}
    current_key: str | None = None
    body_start = 0
    header_seen = False

    aliases = {
        "von": "from", "from": "from", "an": "to", "to": "to", "cc": "cc",
        "betreff": "subject", "subject": "subject", "datum": "date", "date": "date",
        "gesendet": "date", "sent": "date",
    }

    for i, raw_line in enumerate(lines):
        line = raw_line.strip()
        match = HEADER_RE.match(line)
        if match:
            header_seen = True
            current_key = aliases[match.group(1).lower()]
            headers[current_key] = match.group(2).strip()
            body_start = i + 1
            continue

        if header_seen and current_key and line and (raw_line.startswith(" ") or raw_line.startswith("\t")):
            headers[current_key] = (headers.get(current_key, "") + " " + line).strip()
            body_start = i + 1
            continue

        if header_seen:
            body_start = i
            break

    if not header_seen:
        return None

    sender = headers.get("from", "")
    recipients = headers.get("to", "")
    subject = headers.get("subject", "") or Path(source_name).stem
    date_text, time_text = normalize_date(headers.get("date", ""))
    body = clean("\n".join(lines[body_start:]))

    # Mailadressen stehen in Outlook-PDFs oft nur als mailto:-Links.
    mailtos = extract_mailto(reader)
    sender_names = split_people(sender)
    recipient_names = split_people(recipients)
    cc_names = split_people(headers.get("cc", ""))

    sender_address = ""
    recipient_addresses: list[str] = []
    if mailtos:
        sender_address = mailtos[0]
        expected_to_count = len(recipient_names)
        recipient_addresses = mailtos[1:1 + expected_to_count]
        # CC wird bewusst nicht in die Empfaenger-Spalte aufgenommen.
    else:
        sender_found = EMAIL_RE.findall(sender)
        receiver_found = EMAIL_RE.findall(recipients)
        sender_address = sender_found[0] if sender_found else ""
        recipient_addresses = receiver_found

    # Adressen aus den Namensfeldern entfernen.
    sender_display = EMAIL_RE.sub("", sender).strip(" <>;, ")
    recipient_display = EMAIL_RE.sub("", recipients).strip(" <>;, ")

    return {
        "Datum": date_text,
        "Uhrzeit": time_text,
        "Absender": sender_display,
        "Empfaenger": recipient_display,
        "Betreff": subject,
        "Nachrichtentext": body,
        "Absendemailadresse": sender_address,
        "Empfaengermailadresse": "; ".join(recipient_addresses),
    }


def _resolve(obj):
    """Loest ein indirektes PDF-Objekt auf, sofern erforderlich."""
    try:
        return obj.get_object()
    except Exception:
        return obj


def _manual_attachments(reader: PdfReader):
    """Liest eingebettete Dateien direkt aus dem PDF-Namensbaum.

    Dieser Fallback funktioniert auch mit pypdf-Versionen, bei denen
    ``PdfReader.attachments`` noch nicht vorhanden ist.
    """
    try:
        root = _resolve(reader.trailer["/Root"])
        names = _resolve(root.get("/Names"))
        embedded = _resolve(names.get("/EmbeddedFiles")) if names else None
    except Exception:
        embedded = None

    if not embedded:
        return

    def walk_name_tree(node):
        node = _resolve(node)
        if not node:
            return

        pairs = node.get("/Names")
        if pairs:
            pairs = _resolve(pairs)
            for index in range(0, len(pairs) - 1, 2):
                name_obj = _resolve(pairs[index])
                file_spec = _resolve(pairs[index + 1])
                name = str(name_obj)

                if not file_spec:
                    continue

                ef = _resolve(file_spec.get("/EF"))
                if not ef:
                    continue

                stream = ef.get("/F") or ef.get("/UF")
                stream = _resolve(stream)
                if not stream:
                    continue

                try:
                    yield name, stream.get_data()
                except Exception as exc:
                    print(f"Eingebettete Datei konnte nicht gelesen werden: {name}: {exc}")

        for kid in node.get("/Kids", []) or []:
            yield from walk_name_tree(kid)

    yield from walk_name_tree(embedded)


def iter_portfolio_pdfs(portfolio_path: Path):
    reader = PdfReader(str(portfolio_path))
    yielded_hashes: set[int] = set()
    attachment_found = False

    # Neue pypdf-Versionen stellen Attachments komfortabel bereit.
    attachments = getattr(reader, "attachments", None)
    if attachments is not None:
        try:
            for name in attachments:
                clean_name = re.sub(r"^<\d+>", "", str(name))
                if not clean_name.lower().endswith(".pdf"):
                    continue
                for blob in attachments[name]:
                    marker = hash(blob)
                    if marker in yielded_hashes:
                        continue
                    yielded_hashes.add(marker)
                    attachment_found = True
                    try:
                        yield clean_name, PdfReader(BytesIO(blob))
                    except Exception as exc:
                        print(f"Anhang konnte nicht gelesen werden: {clean_name}: {exc}")
        except Exception as exc:
            print(f"Hinweis: pypdf-Anhangsschnittstelle nicht nutzbar ({exc}). Verwende Fallback.")

    # Fallback fuer alte pypdf-Versionen oder unvollstaendige Implementierungen.
    if not attachment_found:
        for name, blob in _manual_attachments(reader) or []:
            clean_name = re.sub(r"^<\d+>", "", str(name))
            if not clean_name.lower().endswith(".pdf"):
                continue
            marker = hash(blob)
            if marker in yielded_hashes:
                continue
            yielded_hashes.add(marker)
            attachment_found = True
            try:
                yield clean_name, PdfReader(BytesIO(blob))
            except Exception as exc:
                print(f"Anhang konnte nicht gelesen werden: {clean_name}: {exc}")

    if not attachment_found:
        yield portfolio_path.name, reader


def write_excel(records: list[dict[str, str]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "E-Mails"
    ws.append(COLUMNS)

    for record in records:
        ws.append([record.get(column, "") for column in COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [13, 11, 30, 42, 55, 100, 38, 55]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def convert(pdf_path: Path) -> tuple[Path, int]:
    records: list[dict[str, str]] = []
    for embedded_name, embedded_reader in iter_portfolio_pdfs(pdf_path):
        try:
            record = parse_email(embedded_reader, embedded_name)
            if record:
                records.append(record)
            else:
                print(f"Keine Outlook-Kopfzeilen erkannt: {embedded_name}")
        except Exception as exc:
            print(f"Fehler in {embedded_name}: {exc}")

    if not records:
        raise RuntimeError("Es wurden keine E-Mails erkannt. Es wird keine leere Excel-Datei erzeugt.")

    output_path = pdf_path.with_name(pdf_path.stem + "_konvertiert.xlsx")
    write_excel(records, output_path)
    return output_path, len(records)


def choose_pdf() -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    filename = filedialog.askopenfilename(
        title="Outlook-PDF oder PDF-Portfolio auswaehlen",
        filetypes=[("PDF-Dateien", "*.pdf"), ("Alle Dateien", "*.*")],
    )
    root.destroy()
    return Path(filename) if filename else None


def main() -> int:
    print("Outlook-PDF-Portfolio-Konverter Version 2.2")
    pdf_path = Path(sys.argv[1]) if len(sys.argv) > 1 else choose_pdf()
    if not pdf_path:
        return 0
    try:
        output_path, count = convert(pdf_path)
        message = f"Fertig.\n\n{count} E-Mails wurden geschrieben.\n{output_path}"
        print(message)
        if len(sys.argv) == 1:
            root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
            messagebox.showinfo("Konvertierung abgeschlossen", message)
            root.destroy()
        return 0
    except Exception as exc:
        print(f"Fehler: {exc}", file=sys.stderr)
        if len(sys.argv) == 1:
            root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
            messagebox.showerror("Fehler", str(exc))
            root.destroy()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
