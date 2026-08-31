#!/usr/bin/env python3
"""CATS-Zeitnachweis (PDF) nach Excel (.xlsx) konvertieren.

Das Programm ist auf das Tabellenlayout des bereitgestellten CATS-Zeitnachweises
zugeschnitten. Es liest den eingebetteten PDF-Text positionsbasiert aus und
benoetigt daher kein OCR.

Aufruf:
    python cats_pdf_to_excel.py eingabe.pdf [-o ausgabe.xlsx]

Ohne Argumente startet eine kleine grafische Oberflaeche (Tkinter).
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

try:
    import pdfplumber
except ImportError as exc:
    raise SystemExit("Fehlendes Paket 'pdfplumber'. Installation: pip install pdfplumber openpyxl") from exc

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit("Fehlendes Paket 'openpyxl'. Installation: pip install pdfplumber openpyxl") from exc

DATE_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
HOURS_RE = re.compile(r"^\d+(?:[.,]\d{2})$")

# Spaltenbereiche in PDF-Punkten. Diese entsprechen dem CATS-Layout.
COLUMN_RANGES = {
    "date": (50, 103),
    "lst": (103, 145),
    "psp": (145, 232),
    "psp_desc": (232, 334),
    "project_title": (334, 446),
    "short_text": (446, 560),
    "hours": (560, 610),
    "status": (610, 700),
    "approver": (700, 820),
}


@dataclass
class Booking:
    arbeitstag: str = ""
    lst_art: str = ""
    psp_element: str = ""
    psp_bezeichnung: str = ""
    techn_projekttitel: str = ""
    kurztext: str = ""
    stunden: float = 0.0
    statustext: str = ""
    genehmiger: str = ""


@dataclass
class Metadata:
    zeitraum: str = ""
    personalnummer: str = ""
    name: str = ""
    kostenstelle: str = ""
    einheit: str = ""
    ma_gruppe: str = ""


def _normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _words_in_range(words: Iterable[dict], x0: float, x1: float) -> str:
    selected = [w for w in words if x0 <= float(w["x0"]) < x1]
    selected.sort(key=lambda w: float(w["x0"]))
    return _normalize_spaces(" ".join(w["text"] for w in selected))


def _group_lines(words: list[dict], tolerance: float = 2.0) -> list[list[dict]]:
    """Gruppiert PDF-Woerter nach ihrer vertikalen Position."""
    lines: list[list[dict]] = []
    for word in sorted(words, key=lambda w: (float(w["top"]), float(w["x0"]))):
        top = float(word["top"])
        if not lines:
            lines.append([word])
            continue
        current_top = sum(float(w["top"]) for w in lines[-1]) / len(lines[-1])
        if abs(top - current_top) <= tolerance:
            lines[-1].append(word)
        else:
            lines.append([word])
    for line in lines:
        line.sort(key=lambda w: float(w["x0"]))
    return lines


def _extract_metadata(first_page) -> Metadata:
    text = first_page.extract_text() or ""
    meta = Metadata()

    m = re.search(r"Arbeitszeitblatt:\s*(\d{2}\.\d{2}\.\d{4}\s+bis\s+\d{2}\.\d{2}\.\d{4})", text)
    if m:
        meta.zeitraum = m.group(1)

    m = re.search(r"Personalnummer:\s*(\d+)\s+([^\n]+)", text)
    if m:
        meta.personalnummer = m.group(1)
        meta.name = _normalize_spaces(m.group(2))

    m = re.search(r"Kostenstelle:\s*([^,\n]+),\s*Einheit:\s*([^\n]*)", text)
    if m:
        meta.kostenstelle = m.group(1).strip()
        meta.einheit = m.group(2).strip()

    m = re.search(r"MA-Gruppe:\s*([^\n]+)", text)
    if m:
        meta.ma_gruppe = m.group(1).strip()

    return meta


def _line_fields(line: list[dict]) -> dict[str, str]:
    return {key: _words_in_range(line, *rng) for key, rng in COLUMN_RANGES.items()}


def _append_continuation(booking: Booking, fields: dict[str, str]) -> None:
    mapping = {
        "psp_desc": "psp_bezeichnung",
        "project_title": "techn_projekttitel",
        "short_text": "kurztext",
        "status": "statustext",
        "approver": "genehmiger",
    }
    for src, dst in mapping.items():
        value = fields[src]
        if value:
            old = getattr(booking, dst)
            setattr(booking, dst, _normalize_spaces(f"{old} {value}" if old else value))


def parse_cats_pdf(pdf_path: str | Path) -> tuple[Metadata, list[Booking], list[tuple[str, float]]]:
    """Liest Buchungen und im PDF ausgewiesene Tages-Summen aus."""
    pdf_path = Path(pdf_path)
    bookings: list[Booking] = []
    day_totals: list[tuple[str, float]] = []
    current: Booking | None = None

    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            raise ValueError("Die PDF-Datei enthaelt keine Seiten.")
        metadata = _extract_metadata(pdf.pages[0])

        for page in pdf.pages:
            words = page.extract_words(x_tolerance=1, y_tolerance=2, keep_blank_chars=False)
            lines = _group_lines(words)

            for line in lines:
                # Nur Tabellenbereich verarbeiten; Kopf/Fuss ignorieren.
                avg_top = sum(float(w["top"]) for w in line) / len(line)
                if avg_top < 195 or avg_top > page.height - 55:
                    continue

                fields = _line_fields(line)
                date = fields["date"]

                if DATE_RE.match(date):
                    has_booking_keys = bool(fields["lst"] and fields["psp"])
                    if has_booking_keys:
                        if current is not None:
                            bookings.append(current)
                        hours = fields["hours"].replace(",", ".")
                        current = Booking(
                            arbeitstag=date,
                            lst_art=fields["lst"],
                            psp_element=fields["psp"],
                            psp_bezeichnung=fields["psp_desc"],
                            techn_projekttitel=fields["project_title"],
                            kurztext=fields["short_text"],
                            stunden=float(hours) if HOURS_RE.match(fields["hours"]) else 0.0,
                            statustext=fields["status"],
                            genehmiger=fields["approver"],
                        )
                    else:
                        # Zeile mit Tages-Summe (Datum + Stunden, ohne LST/PSP).
                        if current is not None:
                            bookings.append(current)
                            current = None
                        if HOURS_RE.match(fields["hours"]):
                            day_totals.append((date, float(fields["hours"].replace(",", "."))))
                elif current is not None:
                    _append_continuation(current, fields)

        if current is not None:
            bookings.append(current)

    if not bookings:
        raise ValueError(
            "Keine CATS-Buchungszeilen erkannt. Moeglicherweise entspricht das PDF nicht dem erwarteten Layout."
        )
    return metadata, bookings, day_totals


def _excel_date(value: str) -> datetime:
    return datetime.strptime(value, "%d.%m.%Y")


def write_excel(
    output_path: str | Path,
    metadata: Metadata,
    bookings: list[Booking],
    day_totals: list[tuple[str, float]],
) -> None:
    output_path = Path(output_path)
    wb = Workbook()

    # Buchungen
    ws = wb.active
    ws.title = "Buchungen"
    headers = [
        "Arbeitstag", "LST-Art", "PSP-Element", "PSP-Bezeichnung",
        "Techn. Projekttitel", "Kurztext", "Stunden", "Statustext",
        "Name des Genehmigers",
    ]
    ws.append(headers)

    for b in bookings:
        ws.append([
            _excel_date(b.arbeitstag), b.lst_art, b.psp_element, b.psp_bezeichnung,
            b.techn_projekttitel, b.kurztext, b.stunden, b.statustext, b.genehmiger,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=2):
        row[0].number_format = "DD.MM.YYYY"
        row[6].number_format = "0.00"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 13, "B": 12, "C": 24, "D": 30, "E": 32,
        "F": 24, "G": 11, "H": 15, "I": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    if len(bookings) >= 1:
        table = Table(displayName="CATS_Buchungen", ref=f"A1:I{len(bookings)+1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)

    # Tagesuebersicht - Werte werden direkt aus dem PDF uebernommen.
    day_ws = wb.create_sheet("Tagesuebersicht")
    day_ws.append(["Arbeitstag", "Stunden laut PDF"])
    for date, hours in day_totals:
        day_ws.append([_excel_date(date), hours])
    for cell in day_ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for row in day_ws.iter_rows(min_row=2):
        row[0].number_format = "DD.MM.YYYY"
        row[1].number_format = "0.00"
    day_ws.column_dimensions["A"].width = 15
    day_ws.column_dimensions["B"].width = 20
    day_ws.freeze_panes = "A2"
    day_ws.sheet_view.showGridLines = False

    # Stammdaten / Dokumentinfos
    info = wb.create_sheet("Info")
    info_rows = [
        ("Zeitraum", metadata.zeitraum),
        ("Personalnummer", metadata.personalnummer),
        ("Name", metadata.name),
        ("Kostenstelle", metadata.kostenstelle),
        ("Einheit", metadata.einheit),
        ("MA-Gruppe", metadata.ma_gruppe),
        ("Anzahl Buchungen", len(bookings)),
        ("Summe Buchungsstunden", sum(b.stunden for b in bookings)),
    ]
    for label, value in info_rows:
        info.append([label, value])
    for cell in info["A"]:
        cell.font = Font(bold=True)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 40
    info["B8"].number_format = "0.00"
    info.sheet_view.showGridLines = False

    wb.save(output_path)


def convert(input_pdf: str | Path, output_xlsx: str | Path | None = None) -> Path:
    input_pdf = Path(input_pdf)
    if output_xlsx is None:
        output_xlsx = input_pdf.with_suffix(".xlsx")
    output_xlsx = Path(output_xlsx)
    metadata, bookings, day_totals = parse_cats_pdf(input_pdf)
    write_excel(output_xlsx, metadata, bookings, day_totals)
    return output_xlsx


def run_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.title("CATS PDF nach Excel")
    root.geometry("560x190")
    root.resizable(False, False)

    input_var = tk.StringVar()
    output_var = tk.StringVar()

    def choose_input() -> None:
        filename = filedialog.askopenfilename(title="CATS-PDF auswaehlen", filetypes=[("PDF", "*.pdf")])
        if filename:
            input_var.set(filename)
            output_var.set(str(Path(filename).with_suffix(".xlsx")))

    def choose_output() -> None:
        filename = filedialog.asksaveasfilename(
            title="Excel-Datei speichern", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")]
        )
        if filename:
            output_var.set(filename)

    def do_convert() -> None:
        if not input_var.get():
            messagebox.showwarning("Hinweis", "Bitte zuerst eine PDF-Datei auswaehlen.")
            return
        try:
            result = convert(input_var.get(), output_var.get() or None)
            messagebox.showinfo("Fertig", f"Excel-Datei wurde erstellt:\n{result}")
        except Exception as exc:  # GUI: Fehler lesbar anzeigen
            messagebox.showerror("Konvertierung fehlgeschlagen", str(exc))

    tk.Label(root, text="CATS-Zeitnachweis (PDF)").grid(row=0, column=0, padx=12, pady=(18, 6), sticky="w")
    tk.Entry(root, textvariable=input_var, width=58).grid(row=1, column=0, padx=12, sticky="w")
    tk.Button(root, text="PDF waehlen ...", command=choose_input, width=16).grid(row=1, column=1, padx=6)

    tk.Label(root, text="Ausgabe (Excel)").grid(row=2, column=0, padx=12, pady=(12, 6), sticky="w")
    tk.Entry(root, textvariable=output_var, width=58).grid(row=3, column=0, padx=12, sticky="w")
    tk.Button(root, text="Speichern als ...", command=choose_output, width=16).grid(row=3, column=1, padx=6)

    tk.Button(root, text="Konvertieren", command=do_convert, width=22).grid(row=4, column=0, padx=12, pady=18, sticky="w")
    root.mainloop()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="CATS-Zeitnachweis-PDF nach Excel konvertieren")
    parser.add_argument("input", nargs="?", help="Eingabe-PDF")
    parser.add_argument("-o", "--output", help="Ausgabe-XLSX")
    args = parser.parse_args(argv)

    if not args.input:
        run_gui()
        return 0

    try:
        result = convert(args.input, args.output)
        print(f"Erstellt: {result}")
        return 0
    except Exception as exc:
        print(f"Fehler: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
