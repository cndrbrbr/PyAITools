#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Konvertiert Exchange-/OVH-Access-Logdateien im JSON-lines-Format nach Excel.

Das Programm verarbeitet alle .log-Dateien in einem ausgewählten Ordner.
Pro .log-Datei wird eine .xlsx-Datei mit gleichem Namen erzeugt.

Beispiel:
    1745921_result_2024-09-20_access.log
wird zu:
    1745921_result_2024-09-20_access.xlsx

Das Programm extrahiert zusätzlich Felder aus:
    _cs-uri-query
    _X-OVH-TO-FREEZE

Installation:
    pip install openpyxl

Start:
    python accesslog_to_xlsx.py

Optional:
    python accesslog_to_xlsx.py "C:\\Pfad\\zu\\Logs"
    python accesslog_to_xlsx.py "C:\\Pfad\\zu\\Logs" --recursive
    python accesslog_to_xlsx.py "C:\\Pfad\\zu\\Logs" --overwrite
"""

import argparse
import csv
import json
import tkinter as tk
from io import StringIO
from pathlib import Path
from tkinter import filedialog
from typing import Any, Dict, List, Set
from urllib.parse import parse_qs, unquote_plus

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXCEL_MAX_ROWS = 1_048_576
DATA_ROWS_PER_SHEET = EXCEL_MAX_ROWS - 1


IMPORTANT_QUERY_FIELDS = [
    "User",
    "DeviceId",
    "DeviceType",
    "Cmd",
    "CorrelationID",
    "cafeReqId",
    "Log",
]


OVH_FREEZE_FIELDS = [
    "freeze_date",
    "freeze_time",
    "freeze_s_ip",
    "freeze_method",
    "freeze_uri_stem",
    "freeze_uri_query",
    "freeze_port",
    "freeze_username",
    "freeze_client_ip",
    "freeze_user_agent",
    "freeze_unknown",
    "freeze_status",
    "freeze_substatus",
    "freeze_win32_status",
    "freeze_time_taken",
]


def choose_directory_with_dialog() -> Path:
    """
    Öffnet einen grafischen Ordner-Auswahldialog.
    """
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    selected_directory = filedialog.askdirectory(
        title="Ordner mit Access-Logdateien auswählen"
    )

    root.destroy()

    if not selected_directory:
        raise SystemExit("Kein Ordner ausgewählt. Programm beendet.")

    return Path(selected_directory).expanduser().resolve()


def safe_excel_value(value: Any) -> Any:
    """
    Bereitet Werte für Excel auf.
    Dicts und Listen werden als JSON gespeichert.
    Formelartige Strings werden entschärft.
    """
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)

    if isinstance(value, str):
        if value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    return value


def read_jsonl_log(path: Path):
    """
    Liest eine JSON-lines-Datei.

    Gibt je Zeile zurück:
        line_no, parsed_dict, raw_line, parse_error
    """
    with path.open("r", encoding="utf-8-sig", errors="replace") as f:
        for line_no, line in enumerate(f, start=1):
            raw = line.rstrip("\n")

            if not raw.strip():
                continue

            try:
                obj = json.loads(raw)

                if isinstance(obj, dict):
                    yield line_no, obj, raw, None
                else:
                    yield line_no, {}, raw, "JSON-Zeile ist kein Objekt"

            except json.JSONDecodeError as e:
                yield line_no, {}, raw, f"JSONDecodeError: {e}"


def parse_query_string(query: str) -> Dict[str, str]:
    """
    Extrahiert Parameter aus _cs-uri-query.

    Beispiel:
        User=peter@example.com&DeviceId=ABC&DeviceType=iPhone&Cmd=Ping

    Rückgabe:
        query_User
        query_DeviceId
        query_DeviceType
        query_Cmd
        query_Log
        usw.
    """
    result: Dict[str, str] = {}

    if not query:
        return result

    cleaned = query.strip()

    if cleaned.startswith("?"):
        cleaned = cleaned[1:]

    parsed = parse_qs(cleaned, keep_blank_values=True)

    for key, values in parsed.items():
        clean_key = key.strip()
        value = ";".join(unquote_plus(v) for v in values)
        result[f"query_{clean_key}"] = value

    # Manche Exchange-Logs enthalten nach Semikolon weitere Werte,
    # z. B. CorrelationID=<empty>;&cafeReqId=...
    if ";" in cleaned:
        parts = cleaned.replace("&", ";").split(";")
        for part in parts:
            if "=" in part:
                key, value = part.split("=", 1)
                key = key.strip()
                value = unquote_plus(value.strip())

                if key:
                    result.setdefault(f"query_{key}", value)

    return result


def parse_exchange_log_blob(log_value: str) -> Dict[str, str]:
    """
    Zerlegt den langen Exchange-ActiveSync-Log-Blob aus query_Log.

    Beispielauszug:
        SC1:1_PrxFrom:172.16.1.15_Ver1:161_HH:ex2.mail.ovh.net_...

    Rückgabe:
        eas_SC1
        eas_PrxFrom
        eas_Ver1
        eas_HH
        eas_SmtpAdrs
        usw.
    """
    result: Dict[str, str] = {}

    if not log_value:
        return result

    decoded = unquote_plus(str(log_value))

    for part in decoded.split("_"):
        if ":" in part:
            key, value = part.split(":", 1)
            key = key.strip()
            value = value.strip()

            if key:
                column = f"eas_{key}"

                # Bei mehrfach vorkommenden Keys Werte sammeln
                if column in result and value:
                    result[column] = result[column] + ";" + value
                else:
                    result[column] = value

    return result


def parse_ovh_to_freeze(value: str) -> Dict[str, str]:
    """
    Zerlegt das Feld _X-OVH-TO-FREEZE.

    Dieses Feld ist eine CSV-artige Originalzeile, z. B.:
        "2024-09-20","00:04:57","172.16.2.20","POST",...

    Die Felder werden als freeze_* Spalten ergänzt.
    """
    result: Dict[str, str] = {}

    if not value:
        return result

    try:
        reader = csv.reader(StringIO(value))
        row = next(reader, [])
    except Exception as e:
        result["freeze_parse_error"] = str(e)
        result["freeze_raw"] = value
        return result

    for idx, field_name in enumerate(OVH_FREEZE_FIELDS):
        if idx < len(row):
            result[field_name] = row[idx]

    if len(row) > len(OVH_FREEZE_FIELDS):
        extra = row[len(OVH_FREEZE_FIELDS):]
        result["freeze_extra_fields"] = json.dumps(extra, ensure_ascii=False)

    # Zusätzlich Query aus dem Freeze-Feld auswerten
    freeze_query = result.get("freeze_uri_query", "")
    if freeze_query:
        freeze_query_fields = parse_query_string(freeze_query)

        for key, val in freeze_query_fields.items():
            result[f"freeze_{key}"] = val

        freeze_log = freeze_query_fields.get("query_Log", "")
        if freeze_log:
            eas_fields = parse_exchange_log_blob(freeze_log)
            for key, val in eas_fields.items():
                result[f"freeze_{key}"] = val

    return result


def enrich_access_log_object(obj: Dict[str, Any]) -> Dict[str, Any]:
    """
    Erweitert einen Logdatensatz um analysierte Felder.
    """
    enriched: Dict[str, Any] = dict(obj)

    query = str(obj.get("_cs-uri-query", "") or "")
    query_fields = parse_query_string(query)
    enriched.update(query_fields)

    log_blob = query_fields.get("query_Log", "")
    if log_blob:
        enriched.update(parse_exchange_log_blob(log_blob))

    freeze_value = str(obj.get("_X-OVH-TO-FREEZE", "") or "")
    freeze_fields = parse_ovh_to_freeze(freeze_value)
    enriched.update(freeze_fields)

    return enriched


def collect_columns(path: Path) -> List[str]:
    """
    Ermittelt alle Spalten aus Original- und Zusatzfeldern.
    """
    keys: Set[str] = set()
    has_errors = False

    for line_no, obj, raw, err in read_jsonl_log(path):
        if err:
            has_errors = True
            continue

        enriched = enrich_access_log_object(obj)
        keys.update(enriched.keys())

    preferred_columns = [
        "_line_no",
        "_eventtime",
        "_date",
        "_time",
        "_EventReceivedTime",
        "_cs-method",
        "_cs-uri-stem",
        "_cs-uri-query",
        "query_User",
        "query_DeviceId",
        "query_DeviceType",
        "query_Cmd",
        "query_CorrelationID",
        "query_cafeReqId",
        "_cs-username",
        "_c-ip",
        "_s-ip",
        "_s-port_num",
        "_sc-status_num",
        "_sc-substatus_num",
        "_sc-win32-status_num",
        "_time-taken_num",
        "_cs-user-agent",
        "_ovh_servername",
        "host",
        "freeze_client_ip",
        "freeze_user_agent",
        "freeze_uri_stem",
        "freeze_status",
        "freeze_time_taken",
        "eas_PrxFrom",
        "eas_HH",
        "eas_SmtpAdrs",
        "eas_Mbx",
        "eas_Cafe",
        "eas_Dc",
        "eas_DevOS",
        "eas_As",
        "eas_TmRcv",
        "eas_TmSt",
        "eas_TmFin",
        "eas_TmCmpl",
    ]

    columns: List[str] = []

    for col in preferred_columns:
        if col == "_line_no" or col in keys:
            columns.append(col)

    for col in sorted(keys):
        if col not in columns:
            columns.append(col)

    if has_errors:
        columns += ["_parse_error", "_raw_line"]

    return columns


def style_worksheet(ws):
    """
    Formatiert ein Excel-Arbeitsblatt.
    """
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in list(column_cells)[:200]:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        width = min(max(max_len + 2, 10), 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def add_pivot_like_summary(wb: Workbook, stats: Dict[str, Dict[str, int]]):
    """
    Fügt einfache Übersichtsblätter hinzu:
    - Befehle / Cmd
    - Benutzer
    - Geräte
    - Client-IPs
    - Statuscodes
    """
    for sheet_name, values in stats.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(["Wert", "Anzahl"])

        for key, count in sorted(values.items(), key=lambda x: (-x[1], str(x[0]))):
            ws.append([key, count])

        style_worksheet(ws)


def increase_counter(stats: Dict[str, Dict[str, int]], category: str, value: Any):
    """
    Zählt Werte für Übersichtsblätter.
    """
    if value is None or value == "":
        value = "(leer)"

    value = str(value)

    if category not in stats:
        stats[category] = {}

    stats[category][value] = stats[category].get(value, 0) + 1


def write_log_to_xlsx(log_path: Path, xlsx_path: Path):
    """
    Schreibt eine einzelne Access-Logdatei in eine Excel-Datei.
    """
    print(f"Verarbeite: {log_path.name}")

    columns = collect_columns(log_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Access_Log_1"
    ws.append(columns)

    sheet_index = 1
    data_row_count_on_sheet = 0
    total_rows = 0
    error_rows = 0

    stats: Dict[str, Dict[str, int]] = {
        "Summary_Cmd": {},
        "Summary_User": {},
        "Summary_Device": {},
        "Summary_ClientIP": {},
        "Summary_Status": {},
        "Summary_UriStem": {},
    }

    for line_no, obj, raw, err in read_jsonl_log(log_path):
        if data_row_count_on_sheet >= DATA_ROWS_PER_SHEET:
            style_worksheet(ws)

            sheet_index += 1
            ws = wb.create_sheet(title=f"Access_Log_{sheet_index}")
            ws.append(columns)
            data_row_count_on_sheet = 0

        if err:
            enriched: Dict[str, Any] = {}
            error_rows += 1
        else:
            enriched = enrich_access_log_object(obj)

        row: List[Any] = []

        for col in columns:
            if col == "_line_no":
                row.append(line_no)
            elif col == "_parse_error":
                row.append(err or "")
            elif col == "_raw_line":
                row.append(raw if err else "")
            else:
                row.append(safe_excel_value(enriched.get(col, "")))

        ws.append(row)

        total_rows += 1
        data_row_count_on_sheet += 1

        if not err:
            increase_counter(stats, "Summary_Cmd", enriched.get("query_Cmd", ""))
            increase_counter(stats, "Summary_User", enriched.get("query_User", enriched.get("_cs-username", "")))
            increase_counter(stats, "Summary_Device", enriched.get("query_DeviceType", ""))
            increase_counter(stats, "Summary_ClientIP", enriched.get("_c-ip", ""))
            increase_counter(stats, "Summary_Status", enriched.get("_sc-status_num", ""))
            increase_counter(stats, "Summary_UriStem", enriched.get("_cs-uri-stem", ""))

    style_worksheet(ws)

    summary = wb.create_sheet(title="Zusammenfassung", index=0)
    summary.append(["Feld", "Wert"])
    summary.append(["Quelldatei", str(log_path)])
    summary.append(["Exceldatei", str(xlsx_path)])
    summary.append(["Logzeilen", total_rows])
    summary.append(["Fehlerhafte JSON-Zeilen", error_rows])
    summary.append(["Anzahl Datenspalten", len(columns)])
    summary.append(["Anzahl Log-Sheets", sheet_index])
    style_worksheet(summary)

    add_pivot_like_summary(wb, stats)

    wb.save(xlsx_path)

    print(f"Fertig: {xlsx_path.name} ({total_rows} Zeilen, {error_rows} Fehler)")


def find_log_files(directory: Path, recursive: bool) -> List[Path]:
    """
    Findet alle .log-Dateien im gewählten Verzeichnis.
    """
    if recursive:
        return sorted(directory.rglob("*.log"))

    return sorted(directory.glob("*.log"))


def main():
    parser = argparse.ArgumentParser(
        description="Wandelt Exchange-/OVH-Access-Logs im JSON-lines-Format in Excel-Dateien um."
    )

    parser.add_argument(
        "directory",
        nargs="?",
        help="Verzeichnis mit den Access-Logdateien. Wenn leer, öffnet sich ein Ordnerdialog."
    )

    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Unterverzeichnisse ebenfalls durchsuchen."
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Bestehende .xlsx-Dateien überschreiben."
    )

    args = parser.parse_args()

    if args.directory:
        directory = Path(args.directory).expanduser().resolve()
    else:
        directory = choose_directory_with_dialog()

    if not directory.exists():
        raise FileNotFoundError(f"Verzeichnis existiert nicht: {directory}")

    if not directory.is_dir():
        raise NotADirectoryError(f"Pfad ist kein Verzeichnis: {directory}")

    log_files = find_log_files(directory, args.recursive)

    if not log_files:
        print("Keine .log-Dateien gefunden.")
        input("Enter drücken zum Beenden...")
        return

    print(f"Ausgewählter Ordner: {directory}")
    print(f"Gefundene Logdateien: {len(log_files)}")
    print()

    for log_path in log_files:
        xlsx_path = log_path.with_suffix(".xlsx")

        if xlsx_path.exists() and not args.overwrite:
            print(f"Überspringe, existiert bereits: {xlsx_path.name}")
            continue

        write_log_to_xlsx(log_path, xlsx_path)

    print()
    print("Alle Dateien verarbeitet.")
    input("Enter drücken zum Beenden...")


if __name__ == "__main__":
    main()