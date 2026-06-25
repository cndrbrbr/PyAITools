#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Konvertiert alle .log-Dateien in einem ausgewählten Verzeichnis nach .xlsx.

Erwartetes Format:
- Eine JSON-Zeile pro Logeintrag
- Beispiel:
  {"_EventReceivedTime":"2024-10-04 15:43:52", ...}

Installation:
    pip install openpyxl

Start:
    python log_to_xlsx.py

Optional auch mit Verzeichnis per Kommandozeile:
    python log_to_xlsx.py "C:\\Pfad\\zu\\Logs"

Optional rekursiv:
    python log_to_xlsx.py "C:\\Pfad\\zu\\Logs" --recursive

Optional bestehende Exceldateien überschreiben:
    python log_to_xlsx.py "C:\\Pfad\\zu\\Logs" --overwrite
"""

import argparse
import json
import tkinter as tk
from pathlib import Path
from tkinter import filedialog
from typing import Any, List, Set

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


EXCEL_MAX_ROWS = 1_048_576
DATA_ROWS_PER_SHEET = EXCEL_MAX_ROWS - 1


def choose_directory_with_dialog() -> Path:
    """
    Öffnet einen grafischen Ordner-Auswahldialog.
    """
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    selected_directory = filedialog.askdirectory(
        title="Ordner mit Logdateien auswählen"
    )

    root.destroy()

    if not selected_directory:
        raise SystemExit("Kein Ordner ausgewählt. Programm beendet.")

    return Path(selected_directory).expanduser().resolve()


def safe_excel_value(value: Any) -> Any:
    """
    Bereitet Werte für Excel auf.
    Listen und Dicts werden als JSON-Text gespeichert.
    Strings mit möglicher Excel-Formel werden entschärft.
    """
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)

    if isinstance(value, str):
        if value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    return value


def read_jsonl_log(path: Path):
    """
    Liest eine JSON-lines Logdatei.

    Gibt je Zeile zurück:
        line_no, parsed_dict, raw_line, parse_error
    """
    with path.open("r", encoding="utf-8-sig", errors="replace") as f:
        for line_no, line in enumerate(f, start=1):
            raw = line.rstrip("\n")

            if not raw.strip():
                continue

            try:
                obj = json.loads(raw)

                if isinstance(obj, dict):
                    yield line_no, obj, raw, None
                else:
                    yield line_no, {}, raw, "JSON-Zeile ist kein Objekt"

            except json.JSONDecodeError as e:
                yield line_no, {}, raw, f"JSONDecodeError: {e}"


def collect_columns(path: Path) -> List[str]:
    """
    Ermittelt alle vorkommenden Spalten aus der Logdatei.
    """
    keys: Set[str] = set()
    has_errors = False

    for line_no, obj, raw, err in read_jsonl_log(path):
        keys.update(obj.keys())

        if err:
            has_errors = True

    columns = ["_line_no"] + sorted(keys)

    if has_errors:
        columns += ["_parse_error", "_raw_line"]

    return columns


def style_worksheet(ws):
    """
    Formatiert ein Arbeitsblatt für bessere Lesbarkeit.
    """
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0

        # Aus Performancegründen nur die ersten 200 Zeilen für die Breite prüfen
        for cell in list(column_cells)[:200]:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        width = min(max(max_len + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def write_log_to_xlsx(log_path: Path, xlsx_path: Path):
    """
    Schreibt eine einzelne .log-Datei in eine .xlsx-Datei.
    Bei sehr großen Logdateien werden mehrere Excel-Blätter erzeugt.
    """
    print(f"Verarbeite: {log_path.name}")

    columns = collect_columns(log_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Log_1"
    ws.append(columns)

    sheet_index = 1
    data_row_count_on_sheet = 0
    total_rows = 0
    error_rows = 0

    for line_no, obj, raw, err in read_jsonl_log(log_path):
        if data_row_count_on_sheet >= DATA_ROWS_PER_SHEET:
            style_worksheet(ws)

            sheet_index += 1
            ws = wb.create_sheet(title=f"Log_{sheet_index}")
            ws.append(columns)
            data_row_count_on_sheet = 0

        row: List[Any] = []

        for col in columns:
            if col == "_line_no":
                row.append(line_no)

            elif col == "_parse_error":
                row.append(err or "")

            elif col == "_raw_line":
                row.append(raw if err else "")

            else:
                row.append(safe_excel_value(obj.get(col, "")))

        ws.append(row)

        total_rows += 1
        data_row_count_on_sheet += 1

        if err:
            error_rows += 1

    style_worksheet(ws)

    summary = wb.create_sheet(title="Zusammenfassung", index=0)
    summary.append(["Feld", "Wert"])
    summary.append(["Quelldatei", str(log_path)])
    summary.append(["Exceldatei", str(xlsx_path)])
    summary.append(["Logzeilen", total_rows])
    summary.append(["Fehlerhafte JSON-Zeilen", error_rows])
    summary.append(["Anzahl Datenspalten", len(columns)])
    summary.append(["Anzahl Log-Sheets", sheet_index])

    style_worksheet(summary)

    wb.save(xlsx_path)

    print(f"Fertig: {xlsx_path.name} ({total_rows} Zeilen, {error_rows} Fehler)")


def find_log_files(directory: Path, recursive: bool) -> List[Path]:
    """
    Findet alle .log-Dateien im gewählten Verzeichnis.
    """
    if recursive:
        return sorted(directory.rglob("*.log"))

    return sorted(directory.glob("*.log"))


def main():
    parser = argparse.ArgumentParser(
        description="Wandelt JSON-lines .log-Dateien in .xlsx-Dateien um."
    )

    parser.add_argument(
        "directory",
        nargs="?",
        help="Verzeichnis mit den Logdateien. Wenn leer, öffnet sich ein Ordnerdialog."
    )

    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Unterverzeichnisse ebenfalls durchsuchen."
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Bestehende .xlsx-Dateien überschreiben."
    )

    args = parser.parse_args()

    if args.directory:
        directory = Path(args.directory).expanduser().resolve()
    else:
        directory = choose_directory_with_dialog()

    if not directory.exists():
        raise FileNotFoundError(f"Verzeichnis existiert nicht: {directory}")

    if not directory.is_dir():
        raise NotADirectoryError(f"Pfad ist kein Verzeichnis: {directory}")

    log_files = find_log_files(directory, args.recursive)

    if not log_files:
        print("Keine .log-Dateien gefunden.")
        input("Enter drücken zum Beenden...")
        return

    print(f"Ausgewählter Ordner: {directory}")
    print(f"Gefundene Logdateien: {len(log_files)}")
    print()

    for log_path in log_files:
        xlsx_path = log_path.with_suffix(".xlsx")

        if xlsx_path.exists() and not args.overwrite:
            print(f"Überspringe, existiert bereits: {xlsx_path.name}")
            continue

        write_log_to_xlsx(log_path, xlsx_path)

    print()
    print("Alle Dateien verarbeitet.")
    input("Enter drücken zum Beenden...")


if __name__ == "__main__":
    main()
